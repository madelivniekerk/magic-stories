"""
Story Magic ✨ — AI-powered storybook creator for kids aged 5–12
Redesigned to match Magic Stories design handoff (June 2026)
"""

import streamlit as st
import streamlit.components.v1 as components
import os, json, re, io, textwrap, requests
from datetime import datetime
from dotenv import load_dotenv
from anthropic import Anthropic
from openai import OpenAI
from openpyxl import Workbook, load_workbook
from reportlab.pdfgen import canvas as rl_canvas
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib.utils import ImageReader

load_dotenv(override=True)
MODEL        = "claude-sonnet-4-6"
BASE_DIR     = os.path.dirname(os.path.abspath(__file__))
STORIES_DIR  = os.path.join(BASE_DIR, "stories")
TRACKER_PATH = os.path.join(BASE_DIR, "storybook_tracker.xlsx")
os.makedirs(STORIES_DIR, exist_ok=True)

st.set_page_config(page_title="Magic Stories ✨", page_icon="📖", layout="wide")

# ── CSS (ported from magic-app.css design handoff) ───────────────────────────
st.markdown("""<style>
@import url('https://fonts.googleapis.com/css2?family=Cinzel:wght@400;600;700&family=Cinzel+Decorative:wght@700;900&family=Fredoka+One&family=Spectral:ital,wght@0,400;0,500;1,400;1,500&family=Nunito:wght@400;600;700;800&display=swap');

:root{
  --bg:#06020f; --bg-soft:#0d0520; --bg-2:#150a30;
  --purple:#7b2fa8; --purple-lt:#c084fc; --purple-deep:#4a1d6e;
  --gold:#d4af37; --gold-lt:#ffd700; --gold-deep:#b8932f;
  --parchment:#f5e6c5; --parchment-2:#efdcb4; --parch-edge:#e2cf9f;
  --ink:#2c1a08; --ink-soft:#6c5022;
  --cream:#e8d5b0; --cream-soft:#bda88a;
  --muted:#9a7ab0;
  --pink:#f9a8d4; --green:#86efac; --lilac:#c084fc; --blue:#93c5fd;
  --line:rgba(212,175,55,0.16); --panel:rgba(20,12,42,0.62);
  --panel-line:rgba(176,111,216,0.26); --shadow:0 26px 54px -30px rgba(0,0,0,.75);
}
html,body,.stApp{
  background:radial-gradient(125% 100% at 50% -8%,#1a0a3e 0%,#0d0520 46%,#06020f 100%) !important;
  font-family:'Nunito',sans-serif !important; color:var(--cream) !important;
}
/* Hide Streamlit's default chrome so the app fills the full page cleanly */
header[data-testid="stHeader"],#MainMenu,footer{ display:none !important; }
.block-container{
  padding-top:2rem !important;
  max-width:1300px !important;
  margin-left:auto !important; margin-right:auto !important;
}
@media(max-width:1280px){
  .block-container{ max-width:960px !important; }
}
@media(max-width:960px){
  .block-container{ max-width:100% !important; padding-left:1.25rem !important; padding-right:1.25rem !important; }
}
h1{font-family:'Fredoka One',cursive !important; font-size:clamp(30px,4.5vw,46px) !important; color:#fff !important; margin:8px 0 0 !important;}
h2{font-family:'Fredoka One',cursive !important; font-size:clamp(28px,3.8vw,40px) !important; color:#fff !important; margin:8px 0 6px !important;}
h3{font-family:'Cinzel',serif !important; color:var(--cream) !important;}
p,li,label{color:var(--cream) !important; font-family:'Nunito',sans-serif !important;}
a{color:var(--purple-lt) !important;}

/* ── appbar ── */
.ms-appbar{
  display:flex; align-items:center; justify-content:space-between;
  padding:14px 0 14px; border-bottom:1.5px solid var(--line);
  margin-bottom:24px;
}
.ms-brand{
  display:flex; align-items:center; gap:10px; cursor:pointer;
}
.ms-brand .spark{color:var(--gold-lt); font-size:22px;}
.ms-brand .name{
  font-family:'Cinzel Decorative',serif; font-weight:700; font-size:clamp(18px,2.2vw,26px);
  background:linear-gradient(180deg,var(--gold-lt),var(--gold-deep));
  -webkit-background-clip:text; background-clip:text; -webkit-text-fill-color:transparent;
}

/* ── builder header ── */
.builder-head{text-align:center; margin-bottom:20px;}
.step-count{font-family:'Cinzel',serif; font-weight:600; letter-spacing:.14em; text-transform:uppercase; font-size:16px; color:var(--gold-lt);}

/* ── progress trail ── */
.trail{display:flex; align-items:center; justify-content:center; gap:4px; margin:16px auto 4px; flex-wrap:wrap; max-width:700px;}
.t-node{display:flex; align-items:center; gap:4px;}
.t-dot{
  width:34px; height:34px; border-radius:50%; display:grid; place-items:center;
  font-family:'Cinzel',serif; font-weight:700; font-size:16px; color:var(--cream-soft);
  background:rgba(20,12,42,.7); border:1.5px solid var(--panel-line); transition:.2s; flex-shrink:0;
}
.t-node.done .t-dot{background:linear-gradient(180deg,var(--gold-lt),var(--gold-deep)); color:#2a1d04; border-color:transparent;}
.t-node.cur  .t-dot{background:linear-gradient(180deg,#8a3fbf,#5e1f8e); color:#fff; border-color:rgba(216,178,87,.6); box-shadow:0 0 0 4px rgba(123,47,168,.22);}
.t-seg{width:12px; height:2px; border-radius:2px; background:var(--panel-line); flex-shrink:0;}
.t-node.done + .t-seg{background:var(--gold-deep);}

/* ── step card ── */
.step-panel{
  background:var(--panel); border:1.5px solid var(--panel-line);
  border-radius:18px; padding:clamp(12px,1.8vw,20px); box-shadow:var(--shadow); margin-top:14px;
}
.eyebrow{
  font-family:'Cinzel',serif; font-weight:400; letter-spacing:.14em;
  text-transform:uppercase; font-size:19px; margin:0 0 6px;
}
.step-prompt{
  font-family:'Spectral',serif; font-style:italic;
  font-size:clamp(22px,2.4vw,27px); color:var(--cream); margin:0 0 20px;
}

/* ── option cards ── */
.opt-card{
  position:relative; display:flex; flex-direction:column; align-items:center;
  justify-content:center; gap:4px; text-align:center; padding:10px 4px 8px;
  border-radius:11px; background:rgba(10,6,24,.6); border:1.5px solid var(--panel-line);
  transition:transform .12s, border-color .15s, background .15s;
}
.opt-card:hover{border-color:var(--purple-lt);}
.opt-card.sel{background:linear-gradient(180deg,rgba(138,63,191,.32),rgba(94,31,142,.24));}
.opt-card .opt-em{font-size:36px; line-height:1;}
.opt-card .opt-lbl{font-weight:700; font-size:14px; line-height:1.2; color:var(--cream);}
.opt-card .opt-check{
  position:absolute; top:3px; right:4px; font-size:14px; font-weight:800;
}

/* ── device rows (magic words) ── */
.dev-row{
  display:flex; align-items:center; gap:14px; padding:14px 16px; border-radius:14px;
  background:rgba(10,6,24,.5); border:1.5px solid var(--panel-line); margin-bottom:8px;
  transition:border-color .15s, background .15s;
}
.dev-row.on{background:rgba(94,31,142,.25); border-color:var(--gold);}
.dev-row.locked{opacity:.5; cursor:not-allowed;}
.dev-row .d-em{font-size:34px; flex-shrink:0;}
.dev-row .d-body{flex:1;}
.dev-row .d-nm{font-family:'Cinzel',serif; font-size:17px; font-weight:600; color:var(--cream);}
.dev-row .d-df{font-family:'Spectral',serif; font-style:italic; font-size:16px; color:var(--cream-soft);}
.dev-row .d-lock{font-size:15px; color:var(--cream-soft);}
.dev-row .d-tg{font-size:22px; color:var(--gold-lt);}

/* ── sentence preview ── */
.dev-preview{
  background:rgba(212,175,55,.06); border:1px solid rgba(212,175,55,.2);
  border-radius:14px; padding:16px 20px; margin-top:16px;
}
.dp-label{font-family:'Cinzel',serif; font-size:15px; letter-spacing:.1em; color:var(--gold); text-transform:uppercase;}
.dp-text{font-family:'Spectral',serif; font-size:19px; color:var(--cream); margin:8px 0 0; line-height:1.7;}
.w-adj{color:#9c3fb0;} .w-adv{color:#b8651f; font-style:italic;}
.w-sim{color:#2f6fbf; font-style:italic;} .w-met{color:#2f7d4a;} .w-per{color:#b0457f; font-style:italic;}

/* ── writing block ── */
.write-block{
  background:rgba(212,175,55,.05); border:1px dashed rgba(212,175,55,.35);
  border-radius:12px; padding:10px 14px; margin-top:12px;
}
.wb-head{font-family:'Spectral',serif; font-style:italic; font-size:18px; color:var(--cream); margin:0 0 12px;}
.starter-chips{display:flex; flex-wrap:wrap; gap:7px; margin-bottom:12px;}
.spark-meter{font-family:'Nunito',sans-serif; font-size:17px; margin:8px 0 0; font-weight:700;}
.spark-meter.lvl0{color:var(--cream-soft);}
.spark-meter.lvl1{color:var(--pink);}
.spark-meter.lvl2{color:var(--gold-lt);}
.spark-meter.lvl3{color:var(--green);}

/* ── recap chips ── */
.recap-wrap{display:flex; flex-wrap:wrap; gap:9px; margin:12px 0;}
.rc-chip{
  background:rgba(212,175,55,.1); border:1px solid rgba(212,175,55,.3);
  border-radius:999px; padding:6px 14px; font-size:17px; font-family:'Nunito',sans-serif;
}
.rc-chip b{color:var(--gold-lt); margin-right:4px;}

/* ── builder nav ── */
.builder-nav{display:flex; justify-content:space-between; align-items:center; margin-top:20px;}

/* ── Streamlit button overrides ── */
.stButton>button{
  font-family:'Cinzel',serif !important; letter-spacing:.04em !important;
  border-radius:12px !important; padding:.55rem 1.4rem !important;
  border:1.5px solid rgba(212,175,55,.4) !important;
  background:linear-gradient(180deg,#8a3fbf,#5e1f8e) !important;
  color:#fdf6e6 !important;
  box-shadow:0 10px 22px -10px rgba(123,47,168,.8) !important;
  transition:all .2s !important; min-height:44px !important;
}
.stButton>button:hover{transform:translateY(-2px) !important; box-shadow:0 14px 28px -10px rgba(123,47,168,.9) !important;}
.stButton>button:disabled{opacity:.35 !important; transform:none !important;}
.stTextArea textarea, .stTextInput input{
  border-radius:12px !important; border:1.5px solid var(--panel-line) !important;
  background:rgba(8,4,18,.7) !important; color:var(--cream) !important;
  font-family:'Nunito',sans-serif !important; font-size:18px !important;
  padding:14px 16px !important;
  transition:min-height .25s ease, font-size .2s ease, box-shadow .2s ease !important;
}
.stTextArea textarea:focus{
  border-color:var(--purple-lt) !important;
  box-shadow:0 0 0 3px rgba(123,47,168,.25) !important;
  min-height:280px !important;
  font-size:21px !important;
}
.stTextInput input:focus{
  border-color:var(--purple-lt) !important;
  box-shadow:0 0 0 3px rgba(123,47,168,.2) !important;
}
.stTextArea textarea::placeholder, .stTextInput input::placeholder{color:var(--cream-soft) !important; opacity:.7 !important;}

/* ── card grid: button hiding handled by JS (see inject_card_btn_js in main) ── */
/* .cgrid-row-marker is injected before each row by render_card_grid */
.cgrid-row-marker{ height:0 !important; overflow:hidden !important; margin:0 !important; padding:0 !important; line-height:0 !important; display:block !important; }

/* ── book spread ── */
@keyframes shimmer{
  0%  {box-shadow:0 0 30px rgba(180,130,255,.3),0 0 60px rgba(180,130,255,.1),-4px 4px 24px rgba(0,0,0,.5);}
  50% {box-shadow:0 0 50px rgba(212,175,55,.4), 0 0 90px rgba(180,130,255,.2),-4px 4px 24px rgba(0,0,0,.5);}
  100%{box-shadow:0 0 30px rgba(180,130,255,.3),0 0 60px rgba(180,130,255,.1),-4px 4px 24px rgba(0,0,0,.5);}
}
@keyframes float{0%,100%{transform:translateY(0);} 50%{transform:translateY(-8px);}}
@keyframes twinkle{0%,100%{opacity:.2;transform:scale(.8);} 50%{opacity:1;transform:scale(1.3);}}
.magic-float{animation:float 4s ease-in-out infinite;}
.twinkle-1{animation:twinkle 2.1s ease-in-out infinite;}
.twinkle-2{animation:twinkle 2.8s ease-in-out infinite .6s;}
.twinkle-3{animation:twinkle 1.9s ease-in-out infinite 1.2s;}
.twinkle-4{animation:twinkle 3.2s ease-in-out infinite .3s;}
.book-outer{position:relative; max-width:980px; margin:0 auto 1.5rem;}
.book-stack-2{position:absolute; top:8px; left:14px; right:14px; bottom:-8px; background:#1a0a2e; border-radius:4px 16px 16px 4px; box-shadow:0 0 15px rgba(180,130,255,.1);}
.book-stack-1{position:absolute; top:4px; left:8px; right:8px; bottom:-4px; background:#22103a; border-radius:4px 16px 16px 4px;}
.book-spread{position:relative; z-index:2; display:flex; min-height:520px; border-radius:4px 16px 16px 4px; overflow:hidden; animation:shimmer 4s ease-in-out infinite;}
.book-spine{width:8px; flex-shrink:0; background:linear-gradient(180deg,#c8a050 0%,#ffd700 15%,#b8901e 30%,#ffd700 50%,#c8a050 70%,#ffd700 85%,#9a7020 100%); box-shadow:-4px 0 12px rgba(0,0,0,.4),4px 0 12px rgba(0,0,0,.3),0 0 8px rgba(212,175,55,.4);}
.book-left{width:50%; position:relative; overflow:hidden; display:flex; align-items:center; justify-content:center; background:#f5e6c8;}
.book-left img{width:100%; height:100%; object-fit:cover; display:block;}
.book-left-magic{width:100%; height:100%; display:flex; flex-direction:column; align-items:center; justify-content:center; padding:2rem; background:radial-gradient(ellipse at center,#2a1050 0%,#12082a 100%);}
.cover-left{width:50%; overflow:hidden; position:relative; display:flex; align-items:center; justify-content:center; background:radial-gradient(ellipse at center,#2a1050,#0d0520);}
.cover-left img{width:100%; height:100%; object-fit:cover; display:block;}
.cover-right{width:50%; display:flex; flex-direction:column; align-items:center; justify-content:center; padding:2.5rem 2rem; text-align:center; background:radial-gradient(ellipse at center,#1e0845,#0a0220);}
.book-right{width:50%; padding:2rem 2.5rem 1.2rem; display:flex; flex-direction:column; position:relative; background:#f5e6c5; background-image:radial-gradient(ellipse at top left,rgba(180,130,255,.06) 0%,transparent 55%),radial-gradient(ellipse at bottom right,rgba(212,175,55,.1) 0%,transparent 55%);}
.book-corner{position:absolute; font-size:1rem; color:#b8912a; opacity:.55;}
.book-running-head{font-family:'Cinzel',serif; font-size:.68rem; color:#9a7030; text-align:center; letter-spacing:.2em; text-transform:uppercase; border-bottom:1px solid rgba(184,145,42,.35); padding-bottom:.5rem; margin-bottom:1.2rem;}
.book-story-text{font-family:'Spectral',serif; font-size:1.1rem; line-height:2; color:#2c1a08; flex:1;}
.book-drop-cap{font-family:'Cinzel Decorative',serif; font-size:3.2rem; font-weight:700; float:left; line-height:.75; margin:.15rem .3rem 0 0; color:#7b2fa8; text-shadow:0 0 12px rgba(123,47,168,.4);}
.book-footer{border-top:1px solid rgba(184,145,42,.35); padding-top:.5rem; margin-top:1rem; text-align:center;}
.book-page-num{font-family:'Cinzel',serif; font-size:.75rem; color:#9a7030; letter-spacing:.1em;}

/* ── end screen ── */
.the-end{
  font-family:'Cinzel Decorative',serif; font-size:clamp(1.8rem,4vw,2.8rem); font-weight:700; text-align:center;
  background:linear-gradient(135deg,var(--gold-lt),var(--purple-lt),var(--gold-lt));
  -webkit-background-clip:text; background-clip:text; -webkit-text-fill-color:transparent;
  filter:drop-shadow(0 0 14px rgba(212,175,55,.4));
}

/* ── library ── */
.lib-card{
  background:rgba(212,175,55,.06); border:1px solid rgba(212,175,55,.2);
  border-radius:18px; padding:1rem 1.3rem; margin-bottom:.8rem;
}
.lib-card .lib-title{font-family:'Cinzel',serif; font-size:1rem; color:var(--gold);}
.lib-card .lib-meta{font-family:'Nunito',sans-serif; font-size:.78rem; color:var(--muted); margin-top:.25rem;}

/* ── info/success overrides ── */
.stAlert{border-radius:14px !important;}
</style>""", unsafe_allow_html=True)

# ── Wizard data (from design handoff magic-app.js) ───────────────────────────
LEVEL_OPTIONS = [
    {"em":"🦄","lbl":"Foundation","sub":"Ages 5–6","val":"short, simple sentences","age":5},
    {"em":"🐉","lbl":"Early Reader","sub":"Ages 7–8","val":"easy, friendly sentences","age":7},
    {"em":"⚡","lbl":"Confident","sub":"Ages 9–10","val":"richer sentences","age":9},
    {"em":"💎","lbl":"Advanced","sub":"Ages 11–12","val":"vivid, layered sentences","age":11},
]
DEVICE_OPTIONS = [
    {"id":"adj", "em":"🎨","nm":"Adjectives",               "df":"Describing words that paint the picture.",                   "minAge":5},
    {"id":"sim", "em":"🪞","nm":"Similes",                  "df":"Compare two things using like or as.",                       "minAge":7},
    {"id":"per", "em":"🌙","nm":"Personification",           "df":"Give feelings or actions to things that aren't alive.",      "minAge":8},
    {"id":"clf", "em":"🎢","nm":"Cliffhangers",              "df":"End a moment on a surprise that makes readers want more.",   "minAge":8},
    {"id":"vso", "em":"🔀","nm":"Varied Sentence Openers",   "df":"Start sentences in different ways to keep writing lively.",  "minAge":8},
    {"id":"met", "em":"🌌","nm":"Metaphors",                 "df":"Say one thing simply is another — no like or as needed.",   "minAge":9},
    {"id":"adv", "em":"🏃","nm":"Adverbs",                  "df":"How something is done — use sparingly for extra punch!",     "minAge":9},
    {"id":"par", "em":"📐","nm":"Paragraphing with Purpose", "df":"Use paragraph breaks to control pace and drama.",            "minAge":9},
    {"id":"fla", "em":"⏪","nm":"Flashback / Memory",        "df":"Step back in time to reveal something important.",           "minAge":10},
    {"id":"rhe", "em":"❓","nm":"Rhetorical Questions",      "df":"Ask a question you don't answer — to make readers think.",  "minAge":10},
    {"id":"for", "em":"🔮","nm":"Foreshadowing",             "df":"Drop a tiny clue about what's coming later in the story.",  "minAge":11},
]
WHO_OPTIONS = [
    {"em":"🦄","lbl":"Brave unicorn","val":"Luna the brave unicorn"},
    {"em":"🦊","lbl":"Curious fox","val":"Pip the curious fox"},
    {"em":"🐉","lbl":"Shy dragon","val":"Ember the shy dragon"},
    {"em":"🧒","lbl":"A hero like me","val":"a brave young explorer just like you"},
    {"em":"🐰","lbl":"Clever rabbit","val":"Sage the clever rabbit"},
    {"em":"🦁","lbl":"Gentle lion","val":"Rumi the gentle lion"},
    {"em":"🐙","lbl":"Octopus inventor","val":"Otto the octopus inventor"},
    {"em":"🦋","lbl":"Brave butterfly","val":"Mira the brave butterfly"},
    {"em":"🤖","lbl":"Friendly robot","val":"Bolt the friendly robot"},
    {"em":"🐻","lbl":"Cuddly bear","val":"Bruno the brave bear"},
    {"em":"🧚","lbl":"Tiny fairy","val":"Faye the tiny fairy"},
    {"em":"🦕","lbl":"Baby dinosaur","val":"Dot the baby dinosaur"},
    {"em":"🐢","lbl":"Wise turtle","val":"Tomo the wise turtle"},
    {"em":"🦉","lbl":"Night owl","val":"Olwen the night owl"},
    {"em":"🐼","lbl":"Panda warrior","val":"Bao the brave panda"},
    {"em":"🦔","lbl":"Tiny hedgehog","val":"Quill the tiny hedgehog"},
    {"em":"🐧","lbl":"Bold penguin","val":"Pim the bold penguin"},
    {"em":"🦝","lbl":"Sneaky raccoon","val":"Roxy the clever raccoon"},
    {"em":"🐝","lbl":"Busy bee","val":"Buzz the busy bee"},
    {"em":"🧜","lbl":"Brave merchild","val":"Marina the brave merchild"},
]
VILLAIN_OPTIONS = [
    {"em":"🐲","lbl":"Dark dragon","val":"Scorch the dark dragon"},
    {"em":"🧙","lbl":"Evil wizard","val":"Maldrix the evil wizard"},
    {"em":"👹","lbl":"Grumpy ogre","val":"Grunk the grumpy ogre"},
    {"em":"🦁","lbl":"Jealous lion","val":"Rex the jealous lion king"},
    {"em":"🕷️","lbl":"Shadow spider","val":"Vex the shadow spider"},
    {"em":"🐍","lbl":"Cunning snake","val":"Sly the cunning snake"},
    {"em":"👿","lbl":"Trickster imp","val":"Nim the trickster imp"},
    {"em":"🌪️","lbl":"Storm witch","val":"Gale the storm witch"},
    {"em":"🦂","lbl":"Desert scorpion","val":"Sting the desert scorpion"},
    {"em":"🐺","lbl":"Lone wolf","val":"Shadow the lone wolf"},
    {"em":"🧟","lbl":"Grumpy zombie","val":"Rott the grumpy zombie"},
    {"em":"👻","lbl":"Mischief ghost","val":"Wisp the mischievous ghost"},
    {"em":"🐊","lbl":"River croc","val":"Snap the river crocodile"},
    {"em":"🦈","lbl":"Storm shark","val":"Razor the storm shark"},
    {"em":"🧛","lbl":"Night count","val":"Vane the count of the night"},
    {"em":"🤖","lbl":"Broken robot","val":"Rust the broken rogue robot"},
    {"em":"🌑","lbl":"Shadow king","val":"Umbra the shadow king"},
    {"em":"🪄","lbl":"Sneaky sorcerer","val":"Trick the sneaky sorcerer"},
    {"em":"🐻","lbl":"Cave bear","val":"Grizzle the cave bear"},
    {"em":"🦉","lbl":"Jealous owl","val":"Dusk the jealous owl sage"},
]
FRIEND_OPTIONS = [
    {"em":"🐶","lbl":"Loyal dog","val":"Rex the loyal dog"},
    {"em":"🐱","lbl":"Magic cat","val":"Whisper the magic cat"},
    {"em":"🦅","lbl":"Eagle scout","val":"Talon the eagle scout"},
    {"em":"🐢","lbl":"Wise turtle","val":"Shell the wise old turtle"},
    {"em":"🦋","lbl":"Fairy guide","val":"Glow the fairy guide"},
    {"em":"🐧","lbl":"Brave penguin","val":"Flip the brave penguin"},
    {"em":"🦜","lbl":"Talking parrot","val":"Pip the talking parrot"},
    {"em":"🐴","lbl":"Winged horse","val":"Skye the winged horse"},
    {"em":"🐸","lbl":"Lucky frog","val":"Leap the lucky frog"},
    {"em":"🐘","lbl":"Strong elephant","val":"Rumble the gentle elephant"},
    {"em":"🐬","lbl":"Dolphin friend","val":"Splash the dolphin"},
    {"em":"🦌","lbl":"Swift deer","val":"Arrow the swift deer"},
    {"em":"🐇","lbl":"Lucky rabbit","val":"Clover the lucky rabbit"},
    {"em":"🌟","lbl":"Star sprite","val":"Lumi the star sprite"},
    {"em":"🧝","lbl":"Wood elf","val":"Leaf the wood elf"},
    {"em":"🐻‍❄️","lbl":"Snow bear","val":"Frost the snow bear"},
    {"em":"🐝","lbl":"Bee companion","val":"Zara the bee companion"},
    {"em":"🧞","lbl":"Friendly genie","val":"Jinn the friendly genie"},
    {"em":"🤖","lbl":"Robot pal","val":"Cog the helpful robot pal"},
    {"em":"🦊","lbl":"Clever helper","val":"Finn the clever helper fox"},
]
WHERE_OPTIONS = [
    {"em":"🌳","lbl":"Enchanted forest","val":"a glowing enchanted forest"},
    {"em":"🏰","lbl":"Sky castle","val":"a castle high in the clouds"},
    {"em":"🌊","lbl":"Under the sea","val":"a kingdom deep beneath the waves"},
    {"em":"🪐","lbl":"Faraway planet","val":"a planet far, far away"},
    {"em":"🏜️","lbl":"Golden desert","val":"a vast golden desert"},
    {"em":"❄️","lbl":"Frozen north","val":"the sparkling frozen north"},
    {"em":"🍄","lbl":"Mushroom village","val":"a tiny mushroom village"},
    {"em":"🌋","lbl":"Dragon mountains","val":"the smoking dragon mountains"},
    {"em":"🏝️","lbl":"Secret island","val":"a hidden secret island"},
    {"em":"🌌","lbl":"Among the stars","val":"high among the twinkling stars"},
    {"em":"🎪","lbl":"Magical circus","val":"a magical travelling circus"},
    {"em":"🕰️","lbl":"Land of clocks","val":"the curious land of clocks"},
    {"em":"🍭","lbl":"Candy kingdom","val":"a sweet kingdom made of candy"},
    {"em":"🏔️","lbl":"Tall mountains","val":"the tallest snowy mountains"},
    {"em":"🦴","lbl":"Dinosaur valley","val":"a valley where dinosaurs still roam"},
    {"em":"📚","lbl":"Living library","val":"a library where the books come alive"},
    {"em":"🌧️","lbl":"Cloud city","val":"a floating city of rain and rainbows"},
    {"em":"🕸️","lbl":"Spooky manor","val":"a friendly but spooky old manor"},
    {"em":"🏞️","lbl":"Whispering valley","val":"a green valley of whispering streams"},
    {"em":"🪺","lbl":"Treetop town","val":"a cosy town built in the treetops"},
]
WHEN_OPTIONS = [
    {"em":"⏳","lbl":"Long ago","val":"long, long ago"},
    {"em":"🌙","lbl":"Moonlit night","val":"on one moonlit night"},
    {"em":"🔮","lbl":"The future","val":"far in the future"},
    {"em":"🌅","lbl":"Summer's day","val":"on the first warm day of summer"},
    {"em":"🍂","lbl":"Autumn dusk","val":"on a crisp autumn evening"},
    {"em":"⛈️","lbl":"A great storm","val":"in the middle of a great storm"},
    {"em":"🎂","lbl":"Their birthday","val":"on their very own birthday"},
    {"em":"🌫️","lbl":"At first light","val":"just as the morning mist lifted"},
    {"em":"🎃","lbl":"Spooky season","val":"on a spooky autumn night"},
    {"em":"🎄","lbl":"Deep midwinter","val":"in the heart of deep midwinter"},
    {"em":"🌸","lbl":"First day of spring","val":"on the very first day of spring"},
    {"em":"☀️","lbl":"Hottest day","val":"on the hottest day of the year"},
    {"em":"🌠","lbl":"Shooting stars","val":"on the night the stars came falling"},
    {"em":"🐓","lbl":"Break of dawn","val":"at the first crow of dawn"},
    {"em":"🌃","lbl":"Stroke of midnight","val":"at the stroke of midnight"},
    {"em":"🏖️","lbl":"Summer holidays","val":"during the long summer holidays"},
    {"em":"🌊","lbl":"Low tide","val":"when the tide slipped far away"},
    {"em":"🪄","lbl":"Blue moon","val":"once in a rare blue moon"},
    {"em":"🌁","lbl":"Festival night","val":"on the night of the great festival"},
    {"em":"🍁","lbl":"Harvest time","val":"at the golden harvest time"},
]
WHAT_OPTIONS = [
    {"em":"⭐","lbl":"Find the star","val":"find the missing star"},
    {"em":"🗝️","lbl":"Secret door","val":"unlock a mysterious secret door"},
    {"em":"🐣","lbl":"Rescue a friend","val":"rescue a lost little friend"},
    {"em":"🧩","lbl":"Solve a riddle","val":"solve an ancient riddle"},
    {"em":"🗺️","lbl":"Find treasure","val":"follow a map to hidden treasure"},
    {"em":"🌈","lbl":"Bring back colour","val":"bring colour back to the world"},
    {"em":"🌱","lbl":"Grow the magic seed","val":"plant and protect the last magic seed"},
    {"em":"🕊️","lbl":"Make peace","val":"make peace between two old rivals"},
    {"em":"🐲","lbl":"Tame a dragon","val":"tame a lonely dragon"},
    {"em":"🎵","lbl":"Find the lost song","val":"find the world's lost song"},
    {"em":"🌟","lbl":"Grant a wish","val":"grant a forgotten wish"},
    {"em":"🌊","lbl":"Save the village","val":"save the village from a great wave"},
    {"em":"🏆","lbl":"Win the great race","val":"win the great race against all odds"},
    {"em":"👑","lbl":"Find the lost crown","val":"find the kingdom's lost crown"},
    {"em":"🔦","lbl":"Explore the cave","val":"explore a deep and mysterious cave"},
    {"em":"🧪","lbl":"Brew the cure","val":"brew the cure for a strange spell"},
    {"em":"🌉","lbl":"Build a bridge","val":"build a bridge to a faraway land"},
    {"em":"🦣","lbl":"Free the giant","val":"free a gentle giant from a trap"},
    {"em":"❄️","lbl":"End the winter","val":"end a winter that would not end"},
    {"em":"🪁","lbl":"Catch the kite","val":"chase a magical runaway kite"},
]
WHY_OPTIONS = [
    {"em":"🌌","lbl":"The sky went dark","val":"the night sky has gone dark"},
    {"em":"❤️","lbl":"A friend needs help","val":"a dear friend is in trouble"},
    {"em":"🤝","lbl":"A big promise","val":"because of a promise they once made"},
    {"em":"✨","lbl":"To find out who","val":"they long to discover who they truly are"},
    {"em":"🏡","lbl":"Find a way home","val":"they are trying to find their way home"},
    {"em":"😊","lbl":"Make someone smile","val":"they want to make someone they love smile again"},
    {"em":"🌟","lbl":"Keep a wish alive","val":"a precious wish is starting to fade"},
    {"em":"🐾","lbl":"Protect the animals","val":"the animals of the land need a champion"},
    {"em":"🌍","lbl":"Save their home","val":"their whole world is in danger"},
    {"em":"🦸","lbl":"Prove they can","val":"everyone said they were too small to try"},
    {"em":"🧩","lbl":"Solve a mystery","val":"a puzzling mystery needs solving"},
    {"em":"💔","lbl":"Mend a friendship","val":"two friends have fallen out"},
    {"em":"🕯️","lbl":"Bring back hope","val":"the people have lost all hope"},
    {"em":"🎁","lbl":"For a special someone","val":"they want to help someone very dear"},
    {"em":"🦋","lbl":"To be brave at last","val":"they are learning to be brave"},
    {"em":"📜","lbl":"Keep a secret safe","val":"an important secret must be protected"},
    {"em":"⭐","lbl":"Follow a dream","val":"they are chasing a long-held dream"},
    {"em":"🌷","lbl":"Make things grow","val":"nothing in the land will grow"},
    {"em":"🕊️","lbl":"Bring back peace","val":"a quarrel has upset the whole land"},
    {"em":"📖","lbl":"Finish a story","val":"an old tale was never finished"},
]

STEPS = [
    {"id":"level","idx":0,"accent":"pink","required":True,
     "eyebrow":"First, a little about the reader",
     "title":"How does your reader read?",
     "prompt":"Pick a reading level — we tune the words, sentences and pace to fit."},
    {"id":"words","idx":1,"accent":"gold","required":False,
     "eyebrow":"Choose your word magic",
     "title":"Pick your magic words",
     "prompt":"Choose writer's tools to practise today. We'll weave them through every page."},
    {"id":"who","idx":2,"accent":"gold","required":True,
     "eyebrow":"Question 1 · Hero","title":"Who is our hero?",
     "prompt":"Every great story needs a hero. Who will ours be today?",
     "options":WHO_OPTIONS,"own_ph":"e.g. a kind robot who loves to bake",
     "write_prompt":"✏️ Tell me more about your hero — what do they look like, and what makes them special?",
     "starters":["My hero looks…","They are really good at…","But they are a little afraid of…"]},
    {"id":"villain","idx":3,"accent":"pink","required":False,
     "eyebrow":"Question 2 · Villain","title":"Who stands in the way?",
     "prompt":"Every great quest has an obstacle. Who (or what) will challenge our hero? (Optional — skip if you prefer!)",
     "options":VILLAIN_OPTIONS,"own_ph":"e.g. a sneaky cloud who steals everyone's shadows",
     "write_prompt":"✏️ What does the villain want, and why are they the way they are?",
     "starters":["The villain wants…","They became this way because…","But deep down, maybe they…"]},
    {"id":"friend","idx":4,"accent":"green","required":False,
     "eyebrow":"Question 3 · Sidekick","title":"Who helps our hero?",
     "prompt":"A loyal companion can make all the difference. Does your hero have a sidekick? (Optional — skip if you prefer!)",
     "options":FRIEND_OPTIONS,"own_ph":"e.g. a tiny glowing snail who always knows the way",
     "write_prompt":"✏️ How did the hero and sidekick meet? What makes them a great team?",
     "starters":["They first met when…","The sidekick's special gift is…","Together they are unstoppable because…"]},
    {"id":"where","idx":5,"accent":"lilac","required":True,
     "eyebrow":"Question 4 · Where","title":"Where does it happen?",
     "prompt":"Now, where in the wide world does our story take place?",
     "options":WHERE_OPTIONS,"own_ph":"e.g. a city built on the back of a whale",
     "write_prompt":"✏️ Paint this place — what would you see, hear, or smell there?",
     "starters":["When you arrive, you can see…","It sounds like…","The air smells of…"]},
    {"id":"when","idx":6,"accent":"purple","required":True,
     "eyebrow":"Question 5 · When","title":"When does our tale unfold?",
     "prompt":"Stories can happen anytime at all. When does ours begin?",
     "options":WHEN_OPTIONS,"own_ph":"e.g. on the night the two moons kissed",
     "write_prompt":"✏️ Set the scene — what is happening in the world at this moment?",
     "starters":["At this time, the world is…","Everyone is busy…","But something feels different…"]},
    {"id":"what","idx":7,"accent":"green","required":True,
     "eyebrow":"Question 6 · What","title":"What's the quest?",
     "prompt":"Here comes the exciting part — what must our hero do?",
     "options":WHAT_OPTIONS,"own_ph":"e.g. teach the grumpy giant to laugh",
     "write_prompt":"✏️ What makes this quest tricky? What might go wrong along the way?",
     "starters":["The hardest part will be…","Along the way, they meet…","To win, my hero must…"]},
    {"id":"why","idx":8,"accent":"pink","required":True,
     "eyebrow":"Question 7 · Why","title":"And why does it matter?",
     "prompt":"The best stories have heart. Why does this quest matter so much?",
     "options":WHY_OPTIONS,
     "write_prompt":"✏️ How does your hero FEEL about this? Why does it matter to them?",
     "starters":["My hero feels…","It matters because…","Deep down, they hope…"]},
    {"id":"touch","idx":9,"accent":"purple","required":False,
     "eyebrow":"The finishing touch","title":"Add your magic touch",
     "prompt":"Anything extra special — a favourite colour, a pet, a funny detail. (Optional.)"},
    {"id":"final","idx":10,"accent":"gold","required":False,
     "eyebrow":"Ready for magic","title":"Your story is ready to be written!",
     "prompt":"Here's everything we'll weave together:"},
]

ACCENT_HEX = {
    "pink":"#f9a8d4","gold":"#ffd700","lilac":"#c084fc",
    "green":"#86efac","purple":"#c084fc","blue":"#93c5fd",
}
AGE_RULES = {
    "Foundation": "VERY simple words (1-2 syllables). Max 6-8 words per sentence. 2-3 sentences per page. Lots of repetition. Fun sounds and simple actions.",
    "Early Reader": "Simple, clear language. 8-12 words per sentence. 3-4 sentences per page. Basic descriptive words and emotions.",
    "Confident": "Varied vocabulary. 10-15 words per sentence. 4-5 sentences per page. Descriptive language, character feelings, simple suspense.",
    "Advanced": "Rich vocabulary. 12-18 words per sentence. 5-6 sentences per page. Complex emotions, moral themes, narrative tension, vivid imagery.",
}
DEVICE_NAMES = {
    "adj":"adjectives (describing words)",
    "sim":"similes (like/as comparisons)",
    "per":"personification (giving feelings to objects)",
    "clf":"cliffhangers (suspenseful endings)",
    "vso":"varied sentence openers",
    "met":"metaphors (saying one thing is another)",
    "adv":"adverbs (how something is done)",
    "par":"paragraphing with purpose",
    "fla":"flashback / memory scenes",
    "rhe":"rhetorical questions",
    "for":"foreshadowing (hinting at what's next)",
}

# ── Backend functions ─────────────────────────────────────────────────────────
def _get_secret(name):
    val = os.getenv(name)
    if val:
        return val
    try:
        return st.secrets[name]
    except Exception:
        return None

def get_client():
    key = _get_secret("ANTHROPIC_API_KEY")
    return Anthropic(api_key=key) if key else None

def get_openai_client():
    key = _get_secret("OPENAI_API_KEY")
    return OpenAI(api_key=key) if key else None


def generate_story(client, wiz):
    level_lbl = wiz.get("level_lbl", "Early Reader")
    level_age = wiz.get("level_age", 7)
    who     = wiz.get("who_val",      "a brave young hero")
    villain = wiz.get("villain_val",  "").strip()
    friend  = wiz.get("friend_val",   "").strip()
    where   = wiz.get("where_val",    "a magical land")
    when    = wiz.get("when_val",     "long ago")
    what    = wiz.get("what_val",     "go on a quest")
    why     = wiz.get("why_val",      "to help someone they love")
    words   = wiz.get("words",        [])
    touch   = wiz.get("touch",        "").strip()
    more    = wiz.get("more",         {})

    age_rules = AGE_RULES.get(level_lbl, AGE_RULES["Early Reader"])

    magic_text = ""
    if words:
        names = [DEVICE_NAMES[w] for w in words if w in DEVICE_NAMES]
        magic_text = f"\nLiterary techniques to weave in: {', '.join(names)}."

    child_lines = []
    for sid, label in [("who","Hero detail"),("villain","Villain detail"),("friend","Sidekick detail"),
                       ("where","Setting detail"),("when","Time detail"),
                       ("what","Quest detail"),("why","Motivation detail")]:
        txt = more.get(sid, "").strip()
        if txt:
            child_lines.append(f'- {label}: "{txt}"')
    child_text = ""
    if child_lines:
        child_text = "\n\nThe child wrote these words — incorporate them verbatim or very closely:\n" + "\n".join(child_lines)

    villain_line = f"- Villain: {villain}" if villain else ""
    friend_line  = f"- Sidekick: {friend}" if friend else ""

    prompt = f"""You are a warm, imaginative children's story writer. Write a beautiful illustrated storybook.

Reading level: {level_lbl} (age {level_age}+)
Language rules: {age_rules}{magic_text}

Story:
- Hero: {who}
{villain_line}
{friend_line}
- Setting: {where}
- Time: {when}
- Quest: {what}
- Motivation: {why}
- Magic touch: {touch or "none"}
{child_text}

Return ONLY valid JSON (no markdown):
{{"title":"...", "pages":[{{"page":1,"text":"..."}}]}}

Rules:
- Exactly 6 pages
- Page 1: introduce hero and world at the given time
- Pages 2–3: the adventure and a challenge{f'; the villain ({villain}) creates the obstacle' if villain else ''}
- Pages 4–5: the hero overcomes it{f'; the sidekick ({friend}) helps at a key moment' if friend else ''}
- Page 6: happy ending and lesson learned
- Keep it positive, fun, and strictly age-appropriate
- Use the hero's name throughout{f'; give the villain a memorable scene' if villain else ''}{f'; give the sidekick a moment to shine' if friend else ''}"""

    resp = client.messages.create(
        model=MODEL, max_tokens=2000,
        messages=[{"role":"user","content":prompt}]
    )
    raw = resp.content[0].text.strip()
    raw = re.sub(r'^```json\s*', '', raw)
    raw = re.sub(r'\s*```$', '', raw)
    return json.loads(raw)


def generate_character_description(character, setting, story_title, client):
    resp = client.messages.create(
        model=MODEL, max_tokens=150,
        messages=[{"role":"user","content":
            f"Write a SHORT (2-3 sentences) visual description of this character for a DALL-E 3 illustration. "
            f"Character: {character}. Story: {story_title}. Setting: {setting}. "
            f"Be VERY specific about exact colours, clothing, distinctive features. "
            f"Keep it compatible with bright watercolour children's book style. Return ONLY the description."}]
    )
    return resp.content[0].text.strip()


def generate_image(page_text, character, setting, page_num, openai_client,
                   custom_hint="", char_description=""):
    style = ("Art style: bright cheerful watercolour children's book illustration, "
             "soft edges, warm pastel palette. No text or letters in the image.")
    char_lock = (f"The main character MUST look exactly like this: {char_description} "
                 if char_description else f"Character: {character}. ")
    if custom_hint:
        prompt = f"{char_lock}Scene: {custom_hint} Setting: {setting}. {style}"
    else:
        prompt = (f"{char_lock}Scene from page {page_num}: {page_text[:220]} "
                  f"Setting: {setting}. {style}")
    response = openai_client.images.generate(
        model="gpt-image-1", prompt=prompt, size="1024x1024", quality="medium", n=1)
    b64 = response.data[0].b64_json
    return f"data:image/png;base64,{b64}"


def save_story_to_tracker(story_id, wiz, story):
    try:
        if os.path.exists(TRACKER_PATH):
            wb = load_workbook(TRACKER_PATH); ws = wb.active
        else:
            wb = Workbook(); ws = wb.active; ws.title = "Story Tracker"
            ws.append(["Story ID","Date","Level","Hero","Setting","When","Quest","Why",
                       "Magic Touch","Title","Page 1","Page 2","Page 3","Page 4","Page 5","Page 6"])
        pages = story.get("pages", [])
        ws.append([story_id, datetime.now().strftime("%Y-%m-%d %H:%M"),
                   wiz.get("level_lbl",""), wiz.get("who_val",""), wiz.get("where_val",""),
                   wiz.get("when_val",""), wiz.get("what_val",""), wiz.get("why_val",""),
                   wiz.get("touch",""), story.get("title","")]
                  + [p.get("text","")[:300] for p in pages])
        wb.save(TRACKER_PATH)
    except Exception:
        pass


def save_story_to_disk(story_id, story, images, wiz):
    story_dir = os.path.join(STORIES_DIR, story_id)
    os.makedirs(story_dir, exist_ok=True)
    image_files = []
    for i, url in enumerate(images):
        if url:
            try:
                import base64 as _b64
                if url.startswith("data:"):
                    img_bytes = _b64.b64decode(url.split(",", 1)[1])
                else:
                    img_bytes = requests.get(url, timeout=30).content
                img_path = os.path.join(story_dir, f"page_{i}.jpg")
                with open(img_path, "wb") as f:
                    f.write(img_bytes)
                image_files.append(f"page_{i}.jpg")
            except Exception:
                image_files.append(None)
        else:
            image_files.append(None)
    data = {
        "story_id": story_id, "date": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "character": wiz.get("who_val",""), "setting": wiz.get("where_val",""),
        "adventure": wiz.get("what_val",""), "extra": wiz.get("touch",""),
        "title": story.get("title",""), "pages": story.get("pages",[]),
        "image_files": image_files,
    }
    with open(os.path.join(story_dir, "story.json"), "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def load_story_library():
    stories = []
    if not os.path.exists(STORIES_DIR):
        return stories
    for entry in os.listdir(STORIES_DIR):
        json_path = os.path.join(STORIES_DIR, entry, "story.json")
        if os.path.exists(json_path):
            try:
                with open(json_path, "r", encoding="utf-8") as f:
                    stories.append(json.load(f))
            except Exception:
                pass
    return sorted(stories, key=lambda x: x.get("date",""), reverse=True)


def generate_story_pdf(story, images, story_id=None):
    buffer = io.BytesIO()
    PW, PH = landscape(A4)
    c = rl_canvas.Canvas(buffer, pagesize=(PW, PH))
    title = story.get("title","My Story")
    pages = story.get("pages",[])
    c.setFillColorRGB(.10,.04,.24); c.rect(0,0,PW,PH,fill=1,stroke=0)
    c.setFillColorRGB(.83,.69,.22); c.setFont("Helvetica-Bold",38)
    words = title.split()
    if len(title)>28 and len(words)>2:
        mid = len(words)//2
        c.drawCentredString(PW/2,PH/2+28," ".join(words[:mid]))
        c.drawCentredString(PW/2,PH/2-8," ".join(words[mid:]))
    else:
        c.drawCentredString(PW/2,PH/2+10,title)
    c.setFillColorRGB(.63,.50,.80); c.setFont("Helvetica-Oblique",16)
    c.drawCentredString(PW/2,PH/2-50,"A Story Just For You")
    c.setFont("Helvetica",11); c.setFillColorRGB(.45,.32,.60)
    c.drawCentredString(PW/2,PH/2-74,"Written with Magic Stories")
    c.showPage()
    for i, page in enumerate(pages):
        img_src = images[i] if i<len(images) else None
        text = page.get("text","")
        if story_id and img_src:
            local_path = os.path.join(STORIES_DIR, story_id, f"page_{i}.jpg")
            if os.path.exists(local_path):
                img_src = local_path
        c.setFillColorRGB(.96,.90,.77); c.rect(0,0,PW,PH,fill=1,stroke=0)
        c.setFillColorRGB(.10,.04,.24); c.rect(0,PH-36,PW,36,fill=1,stroke=0)
        c.setFillColorRGB(.83,.69,.22); c.setFont("Helvetica-Bold",12)
        c.drawCentredString(PW/2,PH-23,title)
        left_w = PW*.50
        if img_src:
            try:
                if img_src.startswith("http"):
                    resp = requests.get(img_src, timeout=20)
                    img_reader = ImageReader(io.BytesIO(resp.content))
                else:
                    img_reader = ImageReader(img_src)
                c.drawImage(img_reader,12,30,left_w-20,PH-80,preserveAspectRatio=True,anchor='c')
            except Exception:
                pass
        c.setStrokeColorRGB(.72,.56,.13); c.setLineWidth(2); c.line(left_w,28,left_w,PH-40)
        tx = left_w+18; tw = PW-left_w-28
        c.setFillColorRGB(.30,.15,.05); c.setFont("Helvetica-Bold",11)
        c.drawString(tx,PH-58,f"Chapter {i+1}")
        c.setFont("Helvetica",11); c.setFillColorRGB(.17,.10,.03)
        wrapped = textwrap.wrap(text, width=int(tw/5.8))
        y = PH-80
        for line in wrapped:
            if y<42: break
            c.drawString(tx,y,line); y-=17
        c.setFillColorRGB(.50,.38,.10); c.setFont("Helvetica-Oblique",9)
        c.drawCentredString(PW/2,16,f"— {i+1} —")
        c.showPage()
    c.save(); buffer.seek(0)
    return buffer.getvalue()


def _fallback_svg(page_num):
    colours = ["#FF6B6B","#4ECDC4","#FFE66D","#A8A4CE","#88D8B0","#FCB69F"]
    col = colours[page_num % len(colours)]
    return (f'<svg viewBox="0 0 500 320" xmlns="http://www.w3.org/2000/svg">'
            f'<rect width="500" height="320" fill="{col}" rx="16"/>'
            f'<text x="250" y="155" font-size="60" text-anchor="middle">📖</text></svg>')


# ── Session state init ────────────────────────────────────────────────────────
_DEFAULTS = {
    "wizard_step":0,
    "wiz_level":None, "wiz_level_age":5,
    "wiz_words":[],
    "wiz_who":None, "wiz_villain":None, "wiz_friend":None,
    "wiz_where":None, "wiz_when":None, "wiz_what":None, "wiz_why":None,
    "wiz_touch":"",
    "more_who":"", "more_villain":"", "more_friend":"",
    "more_where":"", "more_when":"", "more_what":"", "more_why":"",
    "story_data":None, "images":None, "current_page":0, "show_cover":False,
    "char_description":"", "show_library":False,
    "tip_who":"", "tip_villain":"", "tip_friend":"",
    "tip_where":"", "tip_when":"", "tip_what":"", "tip_why":"",
}
for _k, _v in _DEFAULTS.items():
    if _k not in st.session_state:
        st.session_state[_k] = _v


# ── UI Helpers ────────────────────────────────────────────────────────────────
def render_progress_trail():
    total_vis = 10  # 0-9 (not counting final recap)
    cur = st.session_state.wizard_step
    parts = []
    for i in range(total_vis):
        cls = "done" if i < cur else ("cur" if i == cur else "")
        lbl = "✓" if i < cur else str(i + 1)
        parts.append(f'<div class="t-node {cls}"><div class="t-dot">{lbl}</div></div>')
        if i < total_vis - 1:
            parts.append(f'<span class="t-seg"></span>')
    st.markdown('<div class="trail">' + "".join(parts) + "</div>", unsafe_allow_html=True)


def render_card_grid(options, state_key, accent, cols=10, show_sub=False):
    current = st.session_state.get(state_key)
    any_sel = current is not None
    changed = False
    rows = [options[i:i+cols] for i in range(0, len(options), cols)]
    for row in rows:
        st.markdown('<div class="cgrid-row-marker"></div>', unsafe_allow_html=True)
        grid = st.columns(len(row), gap="small")
        for col, opt in zip(grid, row):
            is_sel = any_sel and current.get("val") == opt["val"]

            large = cols <= 4

            if is_sel:
                border   = f"2px solid {accent}"
                bg       = f"linear-gradient(180deg,rgba(138,63,191,.45),rgba(94,31,142,.35))"
                glow     = f"0 0 0 3px {accent}55, 0 4px 24px {accent}33"
                opacity  = "1"
                min_h    = "200px" if large else "140px"
                em_size  = "80px"  if large else "64px"
                lbl_fs   = "18px"  if large else "13px"
                sub_fs   = "15px"  if large else "12px"
                padding  = "28px 10px 24px" if large else "18px 5px 16px"
                lbl_html = f'<span style="font-family:Cinzel,serif;font-weight:400;font-size:{lbl_fs};color:#f0dfc0;letter-spacing:.06em;line-height:1.2;">{opt["lbl"]}</span>'
                check    = f'<span style="position:absolute;top:6px;right:8px;font-size:15px;font-weight:800;color:{accent};">✓</span>'
                sub_html = f'<span style="font-size:{sub_fs};color:#bda88a;font-weight:600;line-height:1.1;">{opt.get("sub","")}</span>' if show_sub and opt.get("sub") else ""
            elif any_sel:
                border   = "1px solid rgba(176,111,216,0.12)"
                bg       = "rgba(10,6,24,.35)"
                glow     = "none"
                opacity  = "0.38"
                min_h    = "80px"  if large else "52px"
                em_size  = "36px"  if large else "26px"
                padding  = "14px 4px 10px" if large else "7px 2px 5px"
                lbl_html = ""
                check    = ""
                sub_html = ""
            else:
                border   = "1.5px solid rgba(176,111,216,0.26)"
                bg       = "rgba(10,6,24,.6)"
                glow     = "none"
                opacity  = "1"
                min_h    = "170px" if large else "88px"
                em_size  = "72px"  if large else "44px"
                lbl_fs   = "17px"  if large else "13px"
                sub_fs   = "14px"  if large else "12px"
                padding  = "24px 10px 20px" if large else "12px 3px 10px"
                lbl_html = f'<span style="font-family:Cinzel,serif;font-weight:400;font-size:{lbl_fs};color:#e8d5b0;letter-spacing:.06em;line-height:1.2;">{opt["lbl"]}</span>'
                check    = ""
                sub_html = f'<span style="font-size:{sub_fs};color:#bda88a;font-weight:600;line-height:1.1;">{opt.get("sub","")}</span>' if show_sub and opt.get("sub") else ""

            col.markdown(
                f'<div style="position:relative;z-index:1;display:flex;flex-direction:column;'
                f'align-items:center;justify-content:center;gap:3px;text-align:center;'
                f'padding:{padding};border-radius:11px;background:{bg};border:{border};'
                f'box-shadow:{glow};min-height:{min_h};opacity:{opacity};">'
                f'{check}'
                f'<span style="font-size:{em_size};line-height:1;">{opt["em"]}</span>'
                f'{lbl_html}{sub_html}</div>',
                unsafe_allow_html=True
            )
            if col.button("⁠", key=f"opt_{state_key}_{opt['val'][:14]}", use_container_width=True):
                st.session_state[state_key] = {"label": opt["lbl"], "val": opt["val"], "custom": False}
                changed = True
    return changed


def render_writing_block(step_id, write_prompt, starters, accent):
    more_key = f"more_{step_id}"
    current_text = st.session_state.get(more_key, "")

    st.markdown(f"""
    <div class="write-block">
      <p class="wb-head">{write_prompt}</p>
    """, unsafe_allow_html=True)

    if st.session_state.get("wiz_words"):
        words_on = st.session_state["wiz_words"]
        names = [DEVICE_NAMES[w].split(" ")[0].capitalize() for w in words_on if w in DEVICE_NAMES]
        if names:
            st.markdown(
                f'<p style="font-family:Cinzel,serif;font-size:11px;color:{accent};'
                f'letter-spacing:.08em;margin:0 0 10px;">✨ Your magic words: {", ".join(names)} '
                f'— try sneaking one in!</p>',
                unsafe_allow_html=True
            )

    if starters:
        st.markdown("**Sentence starters:**", unsafe_allow_html=False)
        s_cols = st.columns(len(starters))
        for sc, starter in zip(s_cols, starters):
            if sc.button(f"＋ {starter}", key=f"start_{step_id}_{starter[:10]}"):
                cur = st.session_state.get(more_key, "")
                st.session_state[more_key] = (cur.rstrip() + " " + starter + " ").lstrip()
                st.rerun()

    text_val = st.text_area(
        "Your writing", value=st.session_state.get(more_key, ""),
        placeholder="Write a sentence or two of your own…",
        key=f"ta_{step_id}", height=160, label_visibility="collapsed"
    )
    st.session_state[more_key] = text_val

    wc = len(text_val.split()) if text_val.strip() else 0
    if wc == 0:
        msg, lvl = "✏️ Your turn — add a sentence or two!", "lvl0"
    elif wc < 5:
        msg, lvl = "✨ Great start — keep going!", "lvl1"
    elif wc < 15:
        msg, lvl = f"✨ {wc} magic words! You're really writing now!", "lvl2"
    else:
        msg, lvl = f"🌟 Wow — {wc} words! What an imagination!", "lvl3"
    st.markdown(f'<p class="spark-meter {lvl}">{msg}</p>', unsafe_allow_html=True)

    # Writing tip button — only show when there's something to react to
    tip_key = f"tip_{step_id}"
    if wc >= 3:
        if st.button("💡 What else could I add?", key=f"tipbtn_{step_id}"):
            child_age = st.session_state.get("wiz_level_age", 7)
            client = get_client()
            if client:
                with st.spinner("Thinking of a question…"):
                    tip = generate_writing_tip(step_id, text_val, child_age, client)
                st.session_state[tip_key] = tip
            else:
                st.session_state[tip_key] = None

    tip = st.session_state.get(tip_key)
    if tip:
        st.markdown(
            f'<div style="margin-top:10px;padding:12px 16px;border-radius:10px;'
            f'background:rgba(123,47,168,.18);border-left:3px solid {accent};">'
            f'<span style="font-family:Nunito,sans-serif;font-size:14px;color:#e8d5b0;">'
            f'✦ {tip}</span></div>',
            unsafe_allow_html=True
        )

    st.markdown('</div>', unsafe_allow_html=True)


def generate_writing_tip(step_id, text, child_age, client):
    step_labels = {
        "who":     "their hero",
        "villain": "the villain",
        "friend":  "the sidekick",
        "where":   "the setting / place",
        "when":    "when the story happens",
        "what":    "the quest or challenge",
        "why":     "why it matters to the hero",
    }
    about = step_labels.get(step_id, "their story")
    prompt = (
        f"A child aged {child_age} is writing about {about} in their storybook. "
        f"They wrote: \"{text.strip()}\"\n\n"
        f"Ask them ONE short, warm, encouraging follow-up question to help them add "
        f"more vivid detail. Keep it to one sentence, age-appropriate for a {child_age}-year-old. "
        f"Don't repeat what they said. Don't start with 'Great!' or similar filler praise. "
        f"Return ONLY the question, nothing else."
    )
    try:
        resp = client.messages.create(
            model=MODEL, max_tokens=80,
            messages=[{"role": "user", "content": prompt}]
        )
        return resp.content[0].text.strip()
    except Exception:
        return None


def render_magic_words(child_age):
    words_on = st.session_state.get("wiz_words", [])

    available = [d for d in DEVICE_OPTIONS if d["minAge"] <= child_age]
    locked    = [d for d in DEVICE_OPTIONS if d["minAge"] > child_age]

    for dev in available:
        is_on = dev["id"] in words_on
        cls   = "dev-row on" if is_on else "dev-row"
        st.markdown(
            f'<div class="{cls}">'
            f'<span class="d-em">{dev["em"]}</span>'
            f'<div class="d-body"><div class="d-nm">{dev["nm"]}</div>'
            f'<div class="d-df">{dev["df"]}</div></div>'
            f'</div>',
            unsafe_allow_html=True
        )
        label = f"{'✓ Remove' if is_on else '+ Add'} {dev['nm']}"
        if st.button(label, key=f"dev_{dev['id']}"):
            lst = list(st.session_state.get("wiz_words", []))
            if dev["id"] in lst:
                lst.remove(dev["id"])
            else:
                lst.append(dev["id"])
            st.session_state["wiz_words"] = lst
            st.rerun()

    if locked:
        # Find which level unlocks the next batch
        next_unlock = min(d["minAge"] for d in locked)
        next_level  = next((o["lbl"] for o in LEVEL_OPTIONS if o["age"] >= next_unlock), None)
        teaser_names = ", ".join(d["nm"] for d in locked[:3])
        more = f" and {len(locked)-3} more" if len(locked) > 3 else ""
        st.markdown(
            f'<div style="margin-top:14px;padding:10px 14px;border-radius:10px;'
            f'background:rgba(123,47,168,.12);border:1px dashed rgba(176,111,216,.25);">'
            f'<span style="font-family:Nunito,sans-serif;font-size:13px;color:#9a7ab0;">'
            f'🔒 {len(locked)} more technique{"s" if len(locked)>1 else ""} unlock at '
            f'<b style="color:#c084fc;">{next_level}</b> level — '
            f'{teaser_names}{more}.'
            f'</span></div>',
            unsafe_allow_html=True
        )

    # Sentence preview
    on = st.session_state.get("wiz_words", [])
    def W(cls, t): return f'<span class="w-{cls}">{t}</span>'
    adj1 = W("adj","emerald ") if "adj" in on else ""
    adj2 = W("adj","rolling ") if "adj" in on else ""
    adv  = W("adv","silently ") if "adv" in on else ""
    s = f"The {adj1}dragon flew {adv}over the {adj2}hills"
    if "sim" in on: s += ", " + W("sim","like a whisper on the wind")
    s += "."
    if "met" in on: s += " " + W("met","The sky was a velvet blanket above.")
    if "per" in on: s += " " + W("per","The moon watched, and the stars whispered back.")

    st.markdown(
        f'<div class="dev-preview"><span class="dp-label">✦ Watch the sentence grow</span>'
        f'<p class="dp-text">{s}</p></div>',
        unsafe_allow_html=True
    )
    st.markdown(
        '<p style="font-family:Spectral,serif;font-style:italic;font-size:13px;color:#bda88a;margin-top:10px;">'
        '✨ Start with a few — more tools unlock as readers grow up.</p>',
        unsafe_allow_html=True
    )


def render_final_recap():
    def chip(k, v):
        if v:
            return f'<div class="rc-chip"><b>{k}</b>{v}</div>'
        return ""

    level     = st.session_state.get("wiz_level")
    who_s     = st.session_state.get("wiz_who")
    villain_s = st.session_state.get("wiz_villain")
    friend_s  = st.session_state.get("wiz_friend")
    where_s   = st.session_state.get("wiz_where")
    when_s    = st.session_state.get("wiz_when")
    what_s    = st.session_state.get("wiz_what")
    why_s     = st.session_state.get("wiz_why")
    words     = st.session_state.get("wiz_words", [])
    touch     = st.session_state.get("wiz_touch","").strip()

    chips = (
        chip("Reader", level.get("label","") if level else "")
        + chip("Hero", who_s.get("val","") if who_s else "")
        + chip("Villain", villain_s.get("val","") if villain_s else "")
        + chip("Sidekick", friend_s.get("val","") if friend_s else "")
        + chip("Where", where_s.get("val","") if where_s else "")
        + chip("When", when_s.get("val","") if when_s else "")
        + chip("What", what_s.get("val","") if what_s else "")
        + chip("Why", why_s.get("val","") if why_s else "")
        + chip("Magic words", ", ".join([DEVICE_NAMES[w].split(" ")[0].capitalize() for w in words]) if words else "")
        + chip("Magic touch", touch)
    )

    # word count from child's writing
    own_words = sum(
        len(st.session_state.get(f"more_{s}","").split())
        for s in ["who","villain","friend","where","when","what","why"]
        if st.session_state.get(f"more_{s}","").strip()
    )

    st.markdown('<div class="recap-wrap">' + chips + '</div>', unsafe_allow_html=True)
    if own_words > 0:
        st.markdown(
            f'<p class="spark-meter lvl3" style="font-size:15px;margin-top:12px;">'
            f'🌟 You wrote {own_words} of your very own words — a true storyteller!</p>',
            unsafe_allow_html=True
        )


def can_advance(step):
    sid = step["id"]
    if not step.get("required", True):
        return True
    if sid == "level":
        return st.session_state.get("wiz_level") is not None
    if sid == "who":
        return st.session_state.get("wiz_who") is not None
    if sid == "where":
        return st.session_state.get("wiz_where") is not None
    if sid == "when":
        return st.session_state.get("wiz_when") is not None
    if sid == "what":
        return st.session_state.get("wiz_what") is not None
    if sid == "why":
        return st.session_state.get("wiz_why") is not None
    return True


# ── Wizard rendering ──────────────────────────────────────────────────────────
def render_step(step):
    sid    = step["id"]
    accent = ACCENT_HEX.get(step["accent"], "#c084fc")

    # Builder header
    step_num = step["idx"] + 1
    total    = len(STEPS) - 1  # exclude final
    label    = "Final flourish" if sid == "final" else f"Step {step_num} of {total}"
    st.markdown(f"""
    <div class="builder-head">
      <div class="step-count">{label}</div>
      <h1>{step["title"]}</h1>
    </div>""", unsafe_allow_html=True)

    render_progress_trail()

    # Step content
    st.markdown(
        f'<div class="step-panel" style="border-color:{accent}33;">'
        f'<p class="eyebrow" style="color:{accent};">{step["eyebrow"]}</p>'
        f'<p class="step-prompt">{step["prompt"]}</p>',
        unsafe_allow_html=True
    )

    if sid == "level":
        changed = render_card_grid(LEVEL_OPTIONS, "wiz_level", accent, cols=4, show_sub=True)
        if changed:
            # Re-read the level age, prune locked magic words
            lvl = st.session_state.get("wiz_level")
            if lvl:
                age = next((o["age"] for o in LEVEL_OPTIONS if o["val"] == lvl.get("val")), 5)
                lvl["age"] = age
                st.session_state["wiz_level_age"] = age
                # Prune words that are now locked
                wds = [w for w in st.session_state.get("wiz_words",[])
                       if any(d["id"]==w and d["minAge"]<=age for d in DEVICE_OPTIONS)]
                st.session_state["wiz_words"] = wds
            st.rerun()
        # Display level label from selection
        sel = st.session_state.get("wiz_level")
        if sel:
            st.success(f"✨ Reading level set to **{sel['label']}**")

    elif sid == "words":
        child_age = st.session_state.get("wiz_level_age", 5)
        render_magic_words(child_age)

    elif sid in ("who","villain","friend","where","when","what","why"):
        options_map = {
            "who":WHO_OPTIONS,"villain":VILLAIN_OPTIONS,"friend":FRIEND_OPTIONS,
            "where":WHERE_OPTIONS,"when":WHEN_OPTIONS,
            "what":WHAT_OPTIONS,"why":WHY_OPTIONS,
        }
        own_ph_map = {
            "who":"e.g. a kind robot who loves to bake",
            "villain":"e.g. a sneaky cloud who steals everyone's shadows",
            "friend":"e.g. a tiny glowing snail who always knows the way",
            "where":"e.g. a city built on the back of a whale",
            "when":"e.g. on the night the two moons kissed",
            "what":"e.g. teach the grumpy giant to laugh",
            "why":"",
        }
        state_key = f"wiz_{sid}"
        opts = options_map[sid]
        changed = render_card_grid(opts, state_key, accent, cols=10)
        if changed:
            st.rerun()

        # Custom / design-your-own
        if step.get("own_ph") or sid != "why":
            with st.expander("✦ …or write your very own"):
                own_val = st.text_input(
                    "Your own choice",
                    placeholder=own_ph_map.get(sid, "Describe your idea…"),
                    key=f"own_{sid}"
                )
                if own_val.strip():
                    st.session_state[state_key] = {"label":"Your own","val":own_val.strip(),"custom":True}

        # Show selection confirmation
        sel = st.session_state.get(state_key)
        if sel:
            st.success(f"✨ {sel['val']}")

        # Writing block
        if step.get("write_prompt"):
            render_writing_block(sid, step["write_prompt"], step.get("starters",[]), accent)

    elif sid == "touch":
        touch_val = st.text_area(
            "Magic touch",
            value=st.session_state.get("wiz_touch",""),
            placeholder="e.g. their best friend is a talking teapot named Doris…",
            key="ta_touch", height=140, label_visibility="collapsed"
        )
        st.session_state["wiz_touch"] = touch_val
        st.markdown(
            '<p style="font-family:Spectral,serif;font-style:italic;font-size:13px;color:#bda88a;margin-top:8px;">'
            'This step is optional — leave it blank if you like.</p>',
            unsafe_allow_html=True
        )

    elif sid == "final":
        render_final_recap()

    st.markdown("</div>", unsafe_allow_html=True)  # close step-panel

    # Navigation
    st.markdown("<div style='height:8px;'></div>", unsafe_allow_html=True)
    nav_l, nav_mid, nav_r = st.columns([1, 2, 1])

    with nav_l:
        if step["idx"] > 0:
            if st.button("← Back", key="nav_back", use_container_width=True):
                st.session_state["wizard_step"] = step["idx"] - 1
                st.session_state["_scroll_top"] = True
                st.rerun()

    with nav_mid:
        if not step.get("required", True) and sid not in ("final","touch"):
            if st.button("Skip this step →", key="nav_skip"):
                st.session_state["wizard_step"] = step["idx"] + 1
                st.session_state["_scroll_top"] = True
                st.rerun()

    with nav_r:
        if sid == "final":
            if st.button("✦  Write My Story!  ✦", key="nav_write", use_container_width=True):
                _build_story(get_client())
        elif step["idx"] < len(STEPS) - 1:
            ok = can_advance(step)
            if not ok:
                st.markdown(
                    '<p style="font-family:Spectral,serif;font-style:italic;font-size:13px;'
                    'color:#f9a8d4;text-align:right;">↑ Make a choice to continue</p>',
                    unsafe_allow_html=True
                )
            next_clicked = st.button("Next →", key="nav_next", use_container_width=True, disabled=not ok)
            if next_clicked:
                st.session_state["wizard_step"] = step["idx"] + 1
                st.session_state["_scroll_top"] = True
                st.rerun()


def _build_story(client):
    if not client:
        st.error("⚠️ ANTHROPIC_API_KEY not found.")
        return

    who   = st.session_state.get("wiz_who")
    where = st.session_state.get("wiz_where")
    level = st.session_state.get("wiz_level")

    villain = st.session_state.get("wiz_villain")
    friend  = st.session_state.get("wiz_friend")
    wiz = {
        "level_lbl":    level.get("label","Early Reader") if level else "Early Reader",
        "level_age":    st.session_state.get("wiz_level_age", 7),
        "who_val":      who.get("val","a brave hero") if who else "a brave hero",
        "villain_val":  villain.get("val","") if villain else "",
        "friend_val":   friend.get("val","") if friend else "",
        "where_val":    where.get("val","a magical land") if where else "a magical land",
        "when_val":     (st.session_state.get("wiz_when") or {}).get("val","long ago"),
        "what_val":     (st.session_state.get("wiz_what") or {}).get("val","go on a quest"),
        "why_val":      (st.session_state.get("wiz_why") or {}).get("val","to help someone they love"),
        "words":        st.session_state.get("wiz_words",[]),
        "touch":        st.session_state.get("wiz_touch",""),
        "more": {s: st.session_state.get(f"more_{s}","")
                 for s in ["who","villain","friend","where","when","what","why"]},
    }

    openai_client = get_openai_client()
    if not openai_client:
        st.warning("⚠️ OPENAI_API_KEY not found — illustrations will be skipped.")
    progress = st.empty()
    bar      = st.progress(0)

    with progress.container():
        st.markdown("""
        <div style="text-align:center;padding:1.5rem 0;">
          <div class="magic-float" style="font-size:3.5rem; margin-bottom:1rem;">📖</div>
          <div style="font-family:'Fredoka One',cursive;font-size:1.6rem;color:#c084fc;">Weaving your story…</div>
          <div style="font-family:'Spectral',serif;font-style:italic;color:#9a7ab0;margin-top:.5rem;">
            The story fairies are dipping their quills…
          </div>
        </div>""", unsafe_allow_html=True)

    bar.progress(8, "Writing your story…")
    try:
        story = generate_story(client, wiz)
    except Exception as e:
        st.error(f"Couldn't write the story: {e}")
        progress.empty(); bar.empty()
        return

    bar.progress(15, "Designing your character…")
    try:
        char_desc = generate_character_description(wiz["who_val"], wiz["where_val"],
                                                   story.get("title",""), client)
    except Exception:
        char_desc = ""
    st.session_state["char_description"] = char_desc

    pages = story.get("pages", [])
    bar.progress(20, "Painting illustrations with DALL-E 3…")
    images = []
    img_errors = []
    for i, page in enumerate(pages):
        pct = 20 + int((i + 1) / len(pages) * 75)
        bar.progress(pct, f"Painting illustration {i+1} of {len(pages)}… 🖌️")
        try:
            url = generate_image(
                page["text"], wiz["who_val"], wiz["where_val"], i+1, openai_client,
                char_description=char_desc
            ) if openai_client else None
        except Exception as img_err:
            img_errors.append(f"Image {i+1}: {img_err}")
            url = None
        images.append(url)

    bar.progress(100, "All done!")
    progress.empty(); bar.empty()

    story_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    save_story_to_tracker(story_id, wiz, story)
    save_story_to_disk(story_id, story, images, wiz)

    st.session_state["story_data"]   = story
    st.session_state["images"]       = images
    st.session_state["img_errors"]   = img_errors
    st.session_state["current_page"] = 0
    st.session_state["show_cover"]   = True
    st.rerun()


# ── Storybook reader ──────────────────────────────────────────────────────────
def show_storybook():
    story  = st.session_state["story_data"]
    images = st.session_state["images"]
    for err in st.session_state.pop("img_errors", []):
        st.error(err)
    pages  = story.get("pages",[])
    title  = story.get("title","My Story")
    total  = len(pages)

    # Cover
    if st.session_state.get("show_cover"):
        cover_img = images[0] if images and images[0] else None
        left_html = (
            f'<img src="{cover_img}" style="width:100%;height:100%;object-fit:cover;display:block;" alt="Cover"/>'
            if cover_img else
            '<div style="display:flex;align-items:center;justify-content:center;width:100%;height:100%;'
            'background:radial-gradient(ellipse at center,#2a1050,#0d0520);">'
            '<span style="font-size:5rem;">✦</span></div>'
        )
        st.markdown(f"""
        <div class="book-outer">
          <div class="book-stack-2"></div><div class="book-stack-1"></div>
          <div class="book-spread">
            <div class="cover-left">{left_html}</div>
            <div class="book-spine"></div>
            <div class="cover-right">
              <div class="twinkle-1" style="font-size:1.3rem;color:#d4af37;letter-spacing:.8rem;margin-bottom:1rem;">✦ ✧ ✦</div>
              <div style="font-family:'Cinzel Decorative',serif;font-size:1.5rem;font-weight:700;line-height:1.35;margin-bottom:1rem;
                   background:linear-gradient(135deg,#ffd700,#f5c842,#ffd700);
                   -webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text;">{title}</div>
              <div style="width:60%;height:1px;background:linear-gradient(90deg,transparent,#d4af37,transparent);margin:0 auto 1rem;"></div>
              <div style="font-family:'Cinzel',serif;font-size:.7rem;color:#9a7ab0;letter-spacing:.22em;text-transform:uppercase;margin-bottom:.4rem;">A Story Just For You</div>
              <div style="font-family:'Cinzel',serif;font-size:.62rem;color:#6a507a;letter-spacing:.18em;text-transform:uppercase;">Written with Magic Stories</div>
              <div class="twinkle-3" style="font-size:1.1rem;color:#d4af37;letter-spacing:.8rem;margin-top:1.2rem;">✧ ✦ ✧</div>
            </div>
          </div>
        </div>""", unsafe_allow_html=True)

        cl, cc, cr = st.columns([1,2,1])
        with cc:
            if st.button("✦  Open the Book  ✦", use_container_width=True):
                st.session_state["show_cover"] = False
                st.session_state["current_page"] = 0
                st.rerun()
        return

    # Story pages
    page_i  = st.session_state["current_page"]
    page    = pages[page_i]
    text    = page.get("text","")
    img_url = images[page_i] if page_i < len(images) else None

    drop = text[0] if text else ""
    rest = text[1:] if len(text) > 1 else ""
    text_html = f'<span class="book-drop-cap">{drop}</span>{rest}' if drop else text

    if img_url:
        left_html = f'<img src="{img_url}" style="width:100%;height:100%;object-fit:cover;display:block;" alt="Illustration"/>'
    else:
        left_html = f"""<div class="book-left-magic">
          <div class="twinkle-2" style="font-size:3rem;margin-bottom:.8rem;">✨</div>
          <div style="font-family:'Cinzel Decorative',serif;font-size:2.4rem;font-weight:700;
               background:linear-gradient(135deg,#ffd700,#c084fc);
               -webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text;margin-bottom:.6rem;">{page_i+1}</div>
          <div class="twinkle-4" style="font-family:serif;font-size:.85rem;color:#d4af37;letter-spacing:.4rem;opacity:.7;">✦ ✧ ✦</div>
          <div style="font-family:'Cinzel',serif;font-size:.62rem;color:#9a7ab0;letter-spacing:.2em;text-transform:uppercase;margin-top:1rem;opacity:.8;">Chapter {page_i+1}</div>
        </div>"""

    st.markdown(f"""
    <div class="book-outer">
      <div class="book-stack-2"></div><div class="book-stack-1"></div>
      <div class="book-spread">
        <div class="book-left">{left_html}</div>
        <div class="book-spine"></div>
        <div class="book-right" style="position:relative;">
          <span class="book-corner" style="top:10px;left:10px;">✦</span>
          <span class="book-corner" style="top:10px;right:10px;">✦</span>
          <span class="book-corner" style="bottom:10px;left:10px;">✦</span>
          <span class="book-corner" style="bottom:10px;right:10px;">✦</span>
          <div class="book-running-head">{title}</div>
          <div class="book-story-text">{text_html}</div>
          <div class="book-footer"><span class="book-page-num">✦ &nbsp; {page_i+1} &nbsp; ✦</span></div>
        </div>
      </div>
    </div>""", unsafe_allow_html=True)

    # Edit page text
    with st.expander("✏️  Edit this page"):
        edited = st.text_area("Edit text", value=text, height=280, key=f"edit_{page_i}", label_visibility="collapsed")
        if st.button("💾  Save changes", key=f"save_{page_i}"):
            st.session_state["story_data"]["pages"][page_i]["text"] = edited
            st.rerun()

    # Page navigation (columns wider than 320px so container query doesn't hide these buttons)
    nl, nm, nr = st.columns([2, 3, 2])
    with nl:
        lbl = "◀  Cover" if page_i == 0 else "◀  Previous"
        if st.button(lbl, use_container_width=True):
            if page_i == 0:
                st.session_state["show_cover"] = True
            else:
                st.session_state["current_page"] -= 1
            st.rerun()
    with nm:
        dots = "".join(
            f'<span style="display:inline-block;width:10px;height:10px;border-radius:50%;margin:0 3px;'
            f'background:{"#d4af37" if i == page_i else "#4a2870"};'
            f'box-shadow:{"0 0 6px rgba(212,175,55,.8)" if i == page_i else "none"};"></span>'
            for i in range(total)
        )
        st.markdown(f'<div style="text-align:center;padding:.6rem 0;">{dots}</div>', unsafe_allow_html=True)
    with nr:
        if page_i < total - 1:
            if st.button("Next page  ▶", use_container_width=True):
                st.session_state["current_page"] += 1
                st.session_state["_scroll_top"] = True
                st.rerun()

    # The End
    if page_i == total - 1:
        st.markdown("""
        <div style="text-align:center;margin:1.5rem 0 .8rem;">
          <div class="the-end magic-float">✦ &nbsp; The End &nbsp; ✦</div>
          <p style="font-family:'Spectral',serif;font-style:italic;color:#9a7ab0;margin-top:.5rem;">
            What a wonderful adventure. Keep it forever, share it tonight, or begin a brand-new tale.
          </p>
        </div>""", unsafe_allow_html=True)

        ca, cb, cc = st.columns(3)
        with ca:
            with st.spinner("📄 Building PDF…"):
                pdf_bytes = generate_story_pdf(story, images)
            st.download_button("📖  Download PDF", data=pdf_bytes,
                               file_name=f"{title.replace(' ','_')}.pdf",
                               mime="application/pdf", use_container_width=True)
        with cb:
            if st.button("📚  Story Library", use_container_width=True):
                st.session_state["show_library"] = True
                st.rerun()
        with cc:
            if st.button("✦  New Story", use_container_width=True):
                for k, v in _DEFAULTS.items():
                    st.session_state[k] = v if not isinstance(v, list) else []
                st.rerun()


# ── Story library ─────────────────────────────────────────────────────────────
def show_library():
    st.markdown("""
    <div style="text-align:center;margin-bottom:1.5rem;">
      <div style="font-family:'Cinzel Decorative',serif;font-size:1.8rem;font-weight:700;
           background:linear-gradient(135deg,#ffd700,#c084fc);
           -webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text;">
        📚 &nbsp; Story Library
      </div>
      <p style="font-family:'Spectral',serif;font-style:italic;color:#9a7ab0;margin-top:.4rem;">
        All your magical adventures — download any story as a PDF anytime.
      </p>
    </div>""", unsafe_allow_html=True)

    stories = load_story_library()
    if not stories:
        st.info("No stories saved yet — write your first one!")
    else:
        for s in stories:
            st.markdown(f"""
            <div class="lib-card">
              <div class="lib-title">{s.get('title','Untitled')}</div>
              <div class="lib-meta">
                📅 {s.get('date','')} &nbsp;·&nbsp;
                🦸 {s.get('character','')} &nbsp;·&nbsp;
                🌍 {s.get('setting','')} &nbsp;·&nbsp;
                ⚔️ {s.get('adventure','')}
              </div>
            </div>""", unsafe_allow_html=True)
            dl_col, _ = st.columns([2, 5])
            with dl_col:
                sid    = s.get("story_id","")
                s_obj  = {"title":s.get("title",""),"pages":s.get("pages",[])}
                imgs   = s.get("image_files",[])
                pdf    = generate_story_pdf(s_obj, imgs, story_id=sid)
                safe   = s.get("title","story").replace(" ","_")[:40]
                st.download_button("📖  Download PDF", data=pdf,
                                   file_name=f"{safe}.pdf", mime="application/pdf",
                                   key=f"dl_{sid}", use_container_width=True)
            st.markdown("<hr style='border:none;border-top:1px solid rgba(212,175,55,.08);margin:.3rem 0 .6rem;'>",
                        unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    if st.button("✦  Write a New Story", use_container_width=False):
        for k, v in _DEFAULTS.items():
            st.session_state[k] = v if not isinstance(v, list) else []
        st.rerun()


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    # JS: walk DOM to find card buttons (after .cgrid-row-marker) and make them invisible+overlaid.
    # CSS selectors proved unreliable across Streamlit versions; direct DOM manipulation is definitive.
    components.html("""<script>
(function(){
  function fix(){
    try{
      var doc=window.parent.document;
      doc.querySelectorAll('.cgrid-row-marker').forEach(function(m){
        // Walk up until we're a direct child of stVerticalBlock
        var el=m;
        while(el.parentElement){
          var pid=el.parentElement.getAttribute('data-testid')||'';
          if(pid.indexOf('VerticalBlock')>=0||pid.indexOf('stVerticalBlock')>=0) break;
          el=el.parentElement;
        }
        var next=el.nextElementSibling;
        if(!next) return;
        next.querySelectorAll('.stButton,[data-testid="stButton"]').forEach(function(bw){
          bw.style.marginTop='-92px';
          bw.style.position='relative';
          bw.style.zIndex='10';
          bw.style.height='0';
          bw.style.minHeight='0';
          bw.style.overflow='visible';
          bw.style.background='transparent';
          bw.style.padding='0';
          var b=bw.querySelector('button');
          if(b){
            b.style.opacity='0';
            b.style.height='90px';
            b.style.minHeight='0';
            b.style.padding='0';
            b.style.border='none';
            b.style.boxShadow='none';
            b.style.cursor='pointer';
            b.style.borderRadius='11px';
            b.style.outline='none';
          }
        });
      });
    }catch(e){}
  }
  fix();
  [80,200,500,1000,2000,3000].forEach(function(t){setTimeout(fix,t);});
  try{
    new MutationObserver(function(){fix();}).observe(
      window.parent.document.body,{childList:true,subtree:true}
    );
  }catch(e){}
})();
</script>""", height=1)

    client = get_client()
    if not client:
        st.error("⚠️ ANTHROPIC_API_KEY not found. Add it to your .env file.")
        return

    # Temporary debug — remove after fixing images
    with st.sidebar:
        st.write("🔑 Anthropic key:", "✅" if client else "❌")
        st.write("🔑 OpenAI key:", "✅" if _get_secret("OPENAI_API_KEY") else "❌")

    # App bar
    acol1, acol2 = st.columns([3, 1])
    with acol1:
        st.markdown("""
        <div class="ms-brand">
          <span class="spark">✦</span>
          <span class="name">Magic Stories</span>
        </div>""", unsafe_allow_html=True)
    with acol2:
        lib_label = "← Builder" if st.session_state.get("show_library") else "📚 Story Library"
        if st.button(lib_label, use_container_width=True):
            st.session_state["show_library"] = not st.session_state.get("show_library", False)
            st.rerun()

    st.markdown("<hr style='border:none;border-top:1.5px solid rgba(212,175,55,.16);margin:.3rem 0 1.2rem;'>",
                unsafe_allow_html=True)

    if st.session_state.pop("_scroll_top", False):
        components.html("""<script>
            (function() {
                var selectors = [
                    '[data-testid="stAppViewBlockContainer"]',
                    '[data-testid="block-container"]',
                    'section.main',
                    '.main',
                    '.stMainBlockContainer'
                ];
                selectors.forEach(function(s) {
                    var el = window.parent.document.querySelector(s);
                    if (el) el.scrollTop = 0;
                });
                window.parent.scrollTo(0, 0);
                window.parent.document.documentElement.scrollTop = 0;
                window.parent.document.body.scrollTop = 0;
            })();
        </script>""", height=0)

    if st.session_state.get("show_library"):
        show_library()
    elif st.session_state.get("story_data") and st.session_state.get("images") is not None:
        show_storybook()
    else:
        step = STEPS[st.session_state.get("wizard_step", 0)]
        render_step(step)


main()
